import ctypes
from ctypes import windll
from platform import platform
import tkinter as ttk
import csv
import os
import sys
import platform
from PIL import Image, ImageTk
from datetime import date
from win32api import GetMonitorInfo, MonitorFromPoint
import customtkinter as ctk
import shutil
import re
from tkcalendar import DateEntry
from datetime import date, datetime
import smtplib
from email.message import EmailMessage
from email.utils import make_msgid
from tkinter import END, INSERT, TclError, filedialog as fd
from openpyxl import load_workbook, workbook

PATH = os.path.dirname(os.path.realpath(__file__))

EMAIL_ADDRESS = ('')
EMAIL_PASSWORD = ('')
gmail_host_smtplib  = 'smtp.gmail.com'

###############################################
################# set scaling #################
###############################################

screensize_old = windll.user32.GetSystemMetrics(0)
windll.shcore.SetProcessDpiAwareness(1)
screensize_new = windll.user32.GetSystemMetrics(0)

scale = round(screensize_old / screensize_new, 2)

ctk.set_window_scaling(scale)
ctk.set_spacing_scaling(scale)
ctk.set_widget_scaling(scale)

###############################################
###############################################

work_area = GetMonitorInfo(MonitorFromPoint((0, 0))).get("Work")
screen_w, screen_h = work_area[2], work_area[3]

root_w = 1200
root_h = 700
border = 20


###############################################
###############################################

class App(ctk.CTk):

    APP_NAME = "Universidad De Manila"
    WIDTH = 1280
    HEIGHT = 768
    image_size_logo = 70
    header_image_size = 450

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title(App.APP_NAME)
        self.geometry(f"{root_w}x{root_h}+{int(screen_w / 2 - root_w / 2)}+{int(screen_h / 2 - root_h / 2)}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.resizable(False, False)
        
        self.server_path = ""

        # load image with PIL and convert to PhotoImage

        image_1_dark = Image.open(PATH + "/image.png").resize((self.WIDTH, self.HEIGHT))
        self.bg_image = ImageTk.PhotoImage(image_1_dark)

        self.image_label = ttk.Label(master=self, image=self.bg_image)
        self.image_label.place(relx=0.5, rely=0.5, anchor=ttk.CENTER)

        self.frame = ctk.CTkFrame(master=self,
                                            width=500,
                                            height=450)
        self.frame.place(relx=0.5, rely=0.5, anchor=ttk.CENTER)

        self.label_1 = ctk.CTkLabel(master=self.frame, width=400, height=150,
                                             text="Universidad De Manila\nTOR Index and \nEmailing App", corner_radius=10,
                                             text_font=('Arial', 15))
        self.label_1.place(relx=0.5, rely=0.3, anchor=ttk.CENTER, x=55)

        # ============ logo_image ============

        image_1 = Image.open(PATH + "/icons/icon.png").resize((App.image_size_logo, App.image_size_logo))
        self.logo_image = ImageTk.PhotoImage(image_1)

        self.image_label = ttk.Label(master=self.frame, image=self.logo_image, borderwidth=0)
        self.image_label.place(relx=0.5, rely=0.5, anchor=ttk.CENTER, x=-100, y=-90)

        # ============ entry_1 & and entry_2 ============

        self.entry_1 = ctk.CTkEntry(master=self.frame, corner_radius=6, width=200, placeholder_text="username")
        self.entry_1.place(relx=0.5, rely=0.52, anchor=ttk.CENTER)
        
        self.entry_2 = ctk.CTkEntry(master=self.frame, corner_radius=6, width=200, show="*", placeholder_text="password")
        self.entry_2.place(relx=0.5, rely=0.6, anchor=ttk.CENTER)

        # ============ button_2 ============

        self.button_2 = ctk.CTkButton(master=self.frame, text="Login",
                                                corner_radius=6, command=self.login, width=200, height=35)
        self.button_2.place(relx=0.5, rely=0.7, anchor=ttk.CENTER)
        
    def login(self):
        username = self.entry_1.get()
        password = self.entry_2.get()
        if username == "admin" and password == "admin":
            self.main_page()
        elif username == "" or password == "" :
            ctypes.windll.user32.MessageBoxW(0, "Please fill up all the blanks ", "Attention", MB_OK | ICON_EXLAIM)
        else:
            ctypes.windll.user32.MessageBoxW(0, "Wrong Account Input details, Try again", "Attention", MB_OK | ICON_STOP)

    def main_page(self):
        for i in self.winfo_children():
            i.destroy()
            
        # ============ create two frames ============
        
        # configure grid layout (2x1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.frame_left = ctk.CTkFrame(master=self,
                                                 width=180,
                                                 corner_radius=0)
        self.frame_left.grid(row=0, column=0, sticky="nswe")

        # ============ frame_right ============

        self.frame_right = ctk.CTkFrame(master=self)
        self.frame_right.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

        # configure grid layout (3x7)
        self.frame_right.rowconfigure((0, 1, 2, 3, 4, 5), weight=1)
        self.frame_right.rowconfigure(10, weight=10)
        self.frame_right.columnconfigure((0, 1, 2, 3), weight=3)
        self.frame_right.columnconfigure(2, weight=0)

        # ============ frame_left ============

        # configure grid layout (1x11)
        self.frame_left.grid_rowconfigure(0, minsize=10)   # empty row with minsize as spacing
        self.frame_left.grid_rowconfigure(8, weight=1)  # empty row as spacing
        self.frame_left.grid_rowconfigure(8, minsize=20)    # empty row with minsize as spacing
        self.frame_left.grid_rowconfigure(11, minsize=10)  # empty row with minsize as spacing

        self.label_1 = ctk.CTkLabel(master=self.frame_left,
                                              text="Universidad De Manila",
                                              text_font=("Roboto Medium", -16))  # font name and size in px
        self.label_1.grid(row=1, column=0, pady=10, padx=10)

        self.button_6 = ctk.CTkButton(master=self.frame_left,
                                                text="TOR Index",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.tor_index,
                                                width=165
                                                )
        self.button_6.grid(row=2, column=0, pady=10, padx=20)

        self.button_7 = ctk.CTkButton(master=self.frame_left,
                                                text="TOR Entry",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.tor_entry,
                                                width=165)
        self.button_7.grid(row=3, column=0, pady=10, padx=20)
        
        self.button_12 = ctk.CTkButton(master=self.frame_left,
                                                text="Appointments",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.notify_request,
                                                width=165)
        self.button_12.grid(row=6, column=0, pady=10, padx=20)
               
        self.button_8 = ctk.CTkButton(master=self.frame_left,
                                                text="Server Check",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.server_status,
                                                width=165)
        self.button_8.grid(row=7, column=0, pady=10, padx=20)
        
        self.button_11 = ctk.CTkButton(master=self.frame_left,
                                                text="Schedule Request",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.notify_request,
                                                width=165)
        self.button_11.grid(row=4, column=0, pady=10, padx=20)
        
        self.button_10 = ctk.CTkButton(master=self.frame_left,
                                                text="Notify Requester",
                                                fg_color=("gray75", "gray30"),  # <- custom tuple-color
                                                command=self.notifying_requester,
                                                width=165)
        self.button_10.grid(row=5, column=0, pady=10, padx=20)
        
        # ============ Header ============
        
        self.label_info_1 = ctk.CTkLabel(master=self.frame_right,
                                                   text="Welcome to Universidad De Manila TOR Index and Emailing App",
                                                   text_font=('Arial, 15'))

        self.label_info_1.grid(column=1, row=0, sticky="nwe", padx=15, pady=15)
        
        image_1_dark = Image.open(PATH + "/icons/icon-background-dark.png").resize((App.header_image_size, App.header_image_size))
        self.bg_image = ImageTk.PhotoImage(image_1_dark)
        self.image_label = ttk.Label(master=self.frame_right, image=self.bg_image, borderwidth=0, bg="#302c2c")
        self.image_label.grid(column=1, row=1, sticky="nwe", padx=0, pady=0)
        
    def tor_index(self):
        self.clear_screen()
        self.fetch_dir()

        courses = [ name for name in os.listdir(self.folder_path) if os.path.isdir(os.path.join(self.folder_path, name)) ]
        self.down_image = self.load_image("/icons/back.png", 60)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="", height=1, width=1, 
                                      corner_radius=10, fg_color="#2a2d2e", hover=False, command=self.back)
        self.button_9.grid(row=0, column=0, pady=(20, 10), sticky="w")
        
        self.label_info_2 = ctk.CTkLabel(master=self.frame_right,
                                         text="TOR Server Index",
                                         text_font=('Arial, 15'))

        self.label_info_2.grid(column=1, row=0, sticky="nwe", padx=15, pady=15)
        
        self.combobox_1 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values=courses, command=self.getUpdateData)
        self.combobox_1.grid(row=1, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_1.entry.configure(readonlybackground="#343638")
        
        self.combobox_2 = ctk.CTkComboBox(master=self.frame_right, state="disabled", command=self.getUpdateData)
        self.combobox_2.grid(row=2, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_2.entry.configure(readonlybackground="#343638")
    
        self.combobox_3 = ctk.CTkComboBox(master=self.frame_right, state="disabled")
        self.combobox_3.grid(row=3, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_3.entry.configure(readonlybackground="#343638")
        
        self.search_image = self.load_image("/icons/open.png", 25)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.search_image, text="Open", height=1, compound="right",
                                      corner_radius=10, command= self.indexed)
        self.button_9.grid(row=5, column=1, pady=(20, 10), sticky="we")
        
    def tor_entry(self):
        self.clear_screen()
        self.fetch_dir()
                    
        courses = [ name for name in os.listdir(self.folder_path) if os.path.isdir(os.path.join(self.folder_path, name)) ] 
        self.down_image = self.load_image("/icons/back.png", 60)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="", height=1, width=1, 
                                      corner_radius=10, fg_color="#2a2d2e", hover=False, command=self.back)
        self.button_9.grid(row=0, column=0, pady=(20, 10), sticky="w")

        self.label_info_2 = ctk.CTkLabel(master=self.frame_right,
                                         text="TOR Server Input",
                                         text_font=('Arial, 15'))

        self.label_info_2.grid(column=1, row=0, sticky="nwe", padx=15, pady=15)
        
        self.combobox_1 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values=courses, command=self.getUpdateData)
        self.combobox_1.grid(row=1, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_1.entry.configure(readonlybackground="#343638")
        
        self.combobox_2 = ctk.CTkComboBox(master=self.frame_right, state="disabled", command=self.getUpdateData)
        self.combobox_2.grid(row=2, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_2.entry.configure(readonlybackground="#343638")
        
        self.entry_1 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               state="readonly")
        self.entry_1.grid(row=3, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.entry_1.entry.configure(readonlybackground="#343638")
        
        self.upload_File = self.load_image("/icons/open.png", 25)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.upload_File, text="Upload", height=1, compound="right",
                                      corner_radius=10, command= self.getFile)
        self.button_9.grid(row=3, column=3, sticky="w")
        
        self.search_image = self.load_image("/icons/open.png", 25)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.search_image, text="Upload to Server", height=1, compound="right",
                                      corner_radius=10, command= self.uploadFile)
        self.button_9.grid(row=4, column=1, pady=(20, 10), sticky="we")
        
    def notify_request(self):
        self.clear_screen()
        self.appointments()
        
        self.down_image = self.load_image("/icons/back.png", 60)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="", height=1, width=1, 
                                      corner_radius=10, fg_color="#2a2d2e", hover=False, command=self.back)
        self.button_9.grid(row=0, column=0, pady=(20, 10), sticky="w")
        
        self.label_info_2 = ctk.CTkLabel(master=self.frame_right,
                                         text="Schedule Request",
                                         text_font=('Arial, 15'))
        self.label_info_2.grid(column=1, row=0, sticky="nwe", padx=15, pady=15)

        self.entry_1 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               placeholder_text= "Name of the Requester")
        self.entry_1.grid(row=2, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.entry_2 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               placeholder_text= "Email")
        self.entry_2.grid(row=3, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.cal = DateEntry(self.frame_right, selectmode = 'day', cursor="hand2",
               year = date.today().year, month = date.today().month,
               day = date.today().day, state="readonly")
        self.cal.grid(row=4, column=1, columnspan=1, pady=10, sticky="we")
        
        time = ['8:00 AM', '9:00 AM', '10:00 AM', '11:00 AM', '1:00 PM', '2:00 PM', '3:00 PM', '4:00 PM']
        for data in range (0, len(time)-1, +1):
            if len(self.data[time[data]]) < 30:
                print("Not yet full")
            else:
                time.pop(data)
        self.combobox_1 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values=time)
        self.combobox_1.grid(row=5, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_1.entry.configure(readonlybackground="#343638")
        self.combobox_1.set('Time Select')
        
        self.schedule_image = self.load_image("/icons/schedule.png", 25)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.schedule_image, text="Schedule", height=1, compound="right",
                                      corner_radius=10, command= self.schedule)
        self.button_9.grid(row=6, column=1, sticky="we")
        
    def notifying_requester(self):
        self.clear_screen()
        self.fetch_dir()
        self.appointments()
        
        courses = [ name for name in os.listdir(self.folder_path) if os.path.isdir(os.path.join(self.folder_path, name)) ]

        self.down_image = self.load_image("/icons/back.png", 60)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="", height=1, width=1, 
                                      corner_radius=10, fg_color="#2a2d2e", hover=False, command=self.back)
        self.button_9.grid(row=0, column=0, pady=(20, 10), sticky="w")
        
        self.label_info_2 = ctk.CTkLabel(master=self.frame_right,
                                         text="Notify Requester",
                                         text_font=('Arial, 15'))

        self.label_info_2.grid(column=1, row=0, sticky="nwe", padx=15, pady=15)
        
        self.combobox_1 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values=courses, command=self.getUpdateData)
        self.combobox_1.grid(row=1, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_1.entry.configure(readonlybackground="#343638")
        self.combobox_1.set('Course Select')
        
        self.combobox_2 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values="", command= self.getUpdateData)
        self.combobox_2.grid(row=2, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_2.entry.configure(readonlybackground="#343638")
        self.combobox_2.set('Section Select')
        
        requests = ['TOR']
        self.combobox_4 = ctk.CTkComboBox(master=self.frame_right, values=requests)
        self.combobox_4.grid(row=3, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_4.entry.configure(readonlybackground="#343638")
        self.combobox_4.set('Request Type Select')

        self.entry_1 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               placeholder_text= "Name of the Student")
        self.entry_1.grid(row=4, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.entry_3 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               placeholder_text= "Student Number")
        self.entry_3.grid(row=5, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.entry_2 = ctk.CTkEntry(master=self.frame_right,
                               width=120,
                               height=28,
                               border_width=2,
                               corner_radius=7,
                               placeholder_text= "Email")
        self.entry_2.grid(row=6, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.cal = DateEntry(self.frame_right, selectmode = 'day', cursor="hand2",
               year = date.today().year, month = date.today().month,
               day = date.today().day)
        self.cal.grid(row=7, column=1, columnspan=1, pady=10, sticky="we")

        time = ['8:00 AM', '9:00 AM', '10:00 AM', '11:00 AM', '1:00 PM', '2:00 PM', '3:00 PM', '4:00 PM']
        for data in range (0, len(time)-1, +1):
            if len(self.data[time[data]]) < 30:
                print("Not yet full")
            else:
                time.pop(data)
        self.combobox_5 = ctk.CTkComboBox(master=self.frame_right, state="readonly", values=time)
        self.combobox_5.grid(row=8, column=1, columnspan=1, pady=10, padx=20, sticky="we")
        self.combobox_5.entry.configure(readonlybackground="#343638")
        self.combobox_5.set('Time Select')
        
        self.notify_image = self.load_image("/icons/notify.png", 25)
        self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.notify_image, text="Notify", height=1, compound="right",
                                      corner_radius=10, command= self.notify)
        self.button_9.grid(row=9, column=1, sticky="we")

    def server_status(self):
        self.clear_screen()
        try:
            os.listdir(self.server_path + "\\Index")
            self.down_image = self.load_image("/icons/check.png", 30)
            self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="Server Up", height=70,
                                        compound="right", text_font="Arial, 20", fg_color="#0066ff", hover=False)
            self.button_9.grid(row=1, column=1, columnspan=1, padx=20, pady=(20, 10), sticky="ew")
        except FileNotFoundError:
            self.down_image = self.load_image("/icons/error.png", 30)
            self.button_9 = ctk.CTkButton(master=self.frame_right, image=self.down_image, text="Server Down", height=70,
                                        compound="right", text_font="Arial, 20", fg_color="#ff0000", hover=False)
            self.button_9.grid(row=1, column=1, columnspan=1, padx=20, pady=(20, 10), sticky="ew")
            
    def appointments(self):
        workbook = self.server_path + "\\Index\\Appointments.xlsx"
        source = load_workbook(workbook)
        date_today = date.today().strftime("%B %d, %Y")
        time = ['8:00 AM', '9:00 AM', '10:00 AM', '11:00 AM', '1:00 PM', '2:00 PM', '3:00 PM', '4:00 PM']
        try:
            sheet_ranges = source[date_today]
        except KeyError:
            print("This Day doesn't exists, Create one")
            source.create_sheet(title=date_today)
            sheet_ranges = source[date_today]
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for iterate in range (0, len(time), +1):
                sheet_ranges[columns[iterate]+ "1"] = time[iterate] 
            source.save(filename=workbook)
        
        self.data = {"8:00 AM": '', "9:00 AM": '', "10:00 AM": '', "11:00 AM": '', "1:00 PM": '', "2:00 PM": '', "3:00 PM": '', 
                "4:00 PM": ''}   
        data_collected = []
        iteration=0
        for columns in sheet_ranges.iter_cols(1, 8, 2, 31, values_only=True):
            for cell in columns:
                if not cell == None:
                    data_collected.append(cell)
            self.data[time[iteration]] = data_collected
            data_collected = []
            iteration+=1
            
    def releasing(self):
        workbook = self.server_path + "\\Index\\Releasing.xlsx"
        source = load_workbook(workbook)
        date_today = date.today().strftime("%B %d, %Y")
        time = ['8:00 AM', '9:00 AM', '10:00 AM', '11:00 AM', '1:00 PM', '2:00 PM', '3:00 PM', '4:00 PM']
        try:
            sheet_ranges = source[date_today]
        except KeyError:
            print("This Day doesn't exists, Create one")
            source.create_sheet(title=date_today)
            sheet_ranges = source[date_today]
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for iterate in range (0, len(time), +1):
                sheet_ranges[columns[iterate]+ "1"] = time[iterate] 
            source.save(filename=workbook)
        
        self.data = {"8:00 AM": '', "9:00 AM": '', "10:00 AM": '', "11:00 AM": '', "1:00 PM": '', "2:00 PM": '', "3:00 PM": '', 
                "4:00 PM": ''}   
        data_collected = []
        iteration=0
        for columns in sheet_ranges.iter_cols(1, 8, 2, 31, values_only=True):
            for cell in columns:
                if not cell == None:
                    data_collected.append(cell)
            self.data[time[iteration]] = data_collected
            data_collected = []
            iteration+=1
    
            
    def getFile(self):
        self.filename = fd.askopenfilename(title="Select a excel file", filetypes=(("Excel Spreadsheet","*.xlsx"),
                                                                              ("Excel 97-2003 spreadsheet","*.xls")))
        self.entry_1.configure(state="normal")
        self.entry_1.insert(0, self.filename)
        self.entry_1.configure(state="readonly")

    def uploadFile(self):
        try:
            filename = os.path.basename(self.filename)
            if not filename == "":
                pass
            else:
                ctypes.windll.user32.MessageBoxW(0, "No file is uploaded, upload a file first", "Attention", MB_OK | ICON_EXLAIM)
                return
        except AttributeError:
            ctypes.windll.user32.MessageBoxW(0, "No file is uploaded, upload a file first", "Attention", MB_OK | ICON_EXLAIM)
            return
        #except UnboundLocalError:
            #ctypes.windll.user32.MessageBoxW(0, "No file is uploaded, upload a file first", "Attention", MB_OK | ICON_EXLAIM)
            #return
        if self.combobox_2.get() == "":
            ctypes.windll.user32.MessageBoxW(0, "Empty Fields found, please fill up all the important details before uploading to server", 
                                             "Attention", MB_OK | ICON_EXLAIM)
            return
        try:
            indexed = [(i, sub.index(os.path.splitext(filename)[0])) for (i, sub) in enumerate(self.index_list) if os.path.splitext(filename)[0] in sub]
        except AttributeError:
            ctypes.windll.user32.MessageBoxW(0, "Empty Fields found, please fill up all the important details before uploading to server", 
                                             "Attention", MB_OK | ICON_EXLAIM)
            return
        if not len(indexed):
            ctypes.windll.user32.MessageBoxW(0, "Filename isn't present in the database, please rename your file to a student number or check the spelling.", 
                                             "Attention", MB_OK | ICON_EXLAIM)
            return
        try:
            shutil.copy2(self.filename, self.folder_path + f"\\{self.course_select}\\{self.section_select}")
        except AttributeError:
            ctypes.windll.user32.MessageBoxW(0, "Empty Fields found, please fill up all the important details before uploading to server", 
                                             "Attention", MB_OK | ICON_EXLAIM)
            return
        
        file_info = re.findall((r'\d+(?:,\d+)?'), str(indexed[0]))
        row_num = int(file_info[0])
        student_num = self.index_list[int(file_info[0])][0]
        file_path = f"{self.server_course}\\{self.course_select}\\{self.section_select}\\{filename}"
        index_file = self.server_path + self.index_file
        with open(index_file, newline="") as f:
            reader = csv.reader(f)
            input_rows = list(reader)

        # Make these row indexes 1-based, so `1: [...]` means "first row"
        override_rows = { row_num: [student_num, file_path] }

        with open(index_file, "w", newline="") as f:
            writer = csv.writer(f)

            # start=0 if your input has header, otherwise start=1
            for row_num, row in enumerate(input_rows, start=0):
                data = override_rows.get(row_num, row)
                writer.writerow(data)
                
        self.create_progress()

    def create_progress(self):
        self.progress_window = ctk.CTkToplevel(self)
        self.progress_window.protocol("WM_DELETE_WINDOW", self.on_closing_progress)
        self.progress_window.title("Uploading TOR")
        self.progress_window.geometry("400x200")
        
        self.label_2 = ctk.CTkLabel(master=self.progress_window,
                                              text="Uploading TOR",
                                              text_font=("Roboto Medium", -16))  # font name and size in px
        self.label_2.pack(pady=20)
        self.value = 0
        self.progressbar = ctk.CTkProgressBar(self.progress_window)
        self.progressbar.pack(pady=20)
        self.progressbar.set(self.value)
        
        self.update_progressbar()

    def update_progressbar(self):
        self.value += 0.01
        try:
            self.progressbar.set(self.value)
        except:
            print("Done Uploading")
            return
        if self.value < 1:
            self.after(20, self.update_progressbar)  # call update_progressbar after 20 ms
        if self.value == 0.9900000000000007:
            self.progress_window.destroy()

    def on_closing(self, event=0):
        self.destroy()
        
    def on_closing_progress(self, event=0):
        self.progress_window.destroy()
        
    def getUpdateData(self, event=0):
        self.course_select = self.combobox_1.get()
        if self.course_select == "Course Select":
            print ("Wait for complete inputs")
            return
        sections = os.listdir(self.folder_path + f"\\{self.course_select}")
        self.combobox_2.configure(values = sections, state="readonly")
        self.section_select = self.combobox_2.get()
        if self.section_select == "Section Select":
            return
        try:
            students = os.listdir(self.folder_path + f"\\{self.course_select}\\{self.section_select}")
        except FileNotFoundError:
            ctypes.windll.user32.MessageBoxW(0, "This section doesn't exists in this course, please choose another section", "Attention", MB_OK | ICON_EXLAIM)
            self.combobox_2.set('')
            try:
                self.combobox_3.set('')
                self.combobox_3.configure(state="disabled")
            except TclError:
                print ("We are in entry, Ignore Combobox3")
            print ("This section doesn't exists")
            
        try:
            if self.section_select:
                try:
                    self.combobox_3.configure(values = students, state="readonly")
                except UnboundLocalError:
                    self.combobox_2.set('')
                    self.combobox_3.set('')
                    self.combobox_3.configure(state="disabled")
                    print ("Unbound Local Error")
        except:
            print("Ignore this errors")

    def start(self):
        self.mainloop()

    def back(self):
        self.clear_screen()
        self.main_page()

    def clear_screen(self):
        for i in self.frame_right.winfo_children():
            i.destroy()

    def load_image(self, path, image_size):
        return ImageTk.PhotoImage(Image.open(PATH + path).resize((image_size, image_size)))
    
    def indexed(self):
        studentFile = self.combobox_3.get()
        try:
            if self.section_select == "" or studentFile == "":
                ctypes.windll.user32.MessageBoxW(0, "Incomplete deatails, Please make sure all fields are field", "Attention", MB_OK | ICON_EXLAIM)
                return
        except AttributeError:
            ctypes.windll.user32.MessageBoxW(0, "Incomplete deatails, Please make sure all fields are field", "Attention", MB_OK | ICON_EXLAIM)
            return  
        os.system(f"start {self.folder_path}\\{self.course_select}\\{self.section_select}\\{studentFile}")
        self.combobox_2.set('')
        self.combobox_3.set('')
        self.combobox_3.configure(state="readonly", values="")
            
    def fetch_dir(self):
        if "" == "":
            self.server_course = "\\Index\\BTVTE"
            self.index_file = "\\Index\\BTVTE_Index.csv"
            with open(self.server_path + self.index_file) as index:
                self.folder_path = self.server_path + self.server_course
                csv_reader = csv.reader(index)
                self.index_list = []
                for row in csv_reader:
                    self.index_list.append(row)
        elif "" == "":
            self.server_course = "\\Index\\CCJ"
            self.index_file = "\\Index\\CCJ_Index.csv"
            with open(self.server_path + self.index_file) as index:
                self.folder_path = self.server_path + self.server_course
                csv_reader = csv.reader(index)
                self.index_list = []
                for row in csv_reader:
                    self.index_list.append(row)
        elif "" == "":
            self.server_course = "\\Index\\BSA"
            self.index_file = "\\Index\\BSA_Index.csv"
            with open(self.server_path + self.index_file) as index:
                self.folder_path = self.server_path + self.server_course
                csv_reader = csv.reader(index)
                self.index_list = []
                for row in csv_reader:
                    self.index_list.append(row)

    def schedule(self):
        self.appointments()
        name = self.entry_1.get()
        email = self.entry_2.get()
        scheduled_datetime = self.cal.get_date().strftime("%B %d, %Y") + " " + self.combobox_1.get()
        regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        if name == "" or name.isspace() or name.startswith(" "):
            ctypes.windll.user32.MessageBoxW(0, "Blank box or space before the text detected, Please fill all boxes", "Attention", MB_OK | ICON_EXLAIM)
            return
        if (re.fullmatch(regex, email)) == None:
            ctypes.windll.user32.MessageBoxW(0, "This is not an email, please check the spelling", "Attention", MB_OK | ICON_EXLAIM)
        if self.combobox_1.get() == "Time Select":
            ctypes.windll.user32.MessageBoxW(0, "Please select desired time first", "Attention", MB_OK | ICON_EXLAIM)
            return
        msg = EmailMessage()
        msg['Subject'] = 'Universidad De Manila - Update for your request'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = email
        prefix = '''<!DOCTYPE html>
        <html lang="en" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:v="urn:schemas-microsoft-com:vml">
        <head>
        <title></title>
        <meta content="text/html; charset=utf-8" http-equiv="Content-Type"/>
        <meta content="width=device-width, initial-scale=1.0" name="viewport"/>
        <!--[if mso]><xml><o:OfficeDocumentSettings><o:PixelsPerInch>96</o:PixelsPerInch><o:AllowPNG/></o:OfficeDocumentSettings></xml><![endif]-->
        <style> 
                * {
                    box-sizing: border-box;
                }
                body {
                    margin: 0;
                    padding: 0;
                }
                a[x-apple-data-detectors] {
                    color: inherit !important;
                    text-decoration: inherit !important;
                }
                #MessageViewBody a {
                    color: inherit;
                    text-decoration: none;
                }
                p {
                    line-height: inherit
                }
                .desktop_hide,
                .desktop_hide table {
                    mso-hide: all;
                    display: none;
                    max-height: 0px;
                    overflow: hidden;
                }
                @media (max-width:620px) {
                    .desktop_hide table.icons-inner {
                        display: inline-block !important;
                    }
                    .icons-inner {
                        text-align: center;
                    }
                    .icons-inner td {
                        margin: 0 auto;
                    }
                    .fullMobileWidth,
                    .image_block img.big,
                    .row-content {
                        width: 100% !important;
                    }
                    .mobile_hide {
                        display: none;
                    }
                    .stack .column {
                        width: 100%;
                        display: block;
                    }
                    .mobile_hide {
                        min-height: 0;
                        max-height: 0;
                        max-width: 0;
                        overflow: hidden;
                        font-size: 0px;
                    }
                    .desktop_hide,
                    .desktop_hide table {
                        display: table !important;
                        max-height: none !important;
                    }
                }
            </style>
        </head>
        <body style="background-color: #FFFFFF; margin: 0; padding: 0; -webkit-text-size-adjust: none; text-size-adjust: none;">
        <table border="0" cellpadding="0" cellspacing="0" class="nl-container" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #FFFFFF;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; background-image: url('https://i.ibb.co/LptDwSH/blue-glow-3-2.png'); background-repeat: no-repeat; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:35px;padding-left:30px;padding-right:30px;padding-top:35px;width:100%;">
        <div align="center" class="alignment" style="line-height:10px"><img src="https://i.ibb.co/JtmhRLG/icon.png" style="display: block; height: auto; border: 0; width: 150px; max-width: 100%;" width="150"/></div>
        </td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="width:100%;padding-right:0px;padding-left:0px;">
        <div align="center" class="alignment" style="line-height:10px"><img class="fullMobileWidth" src="https://d1oco4z2z1fhwp.cloudfront.net/templates/default/4011/top-rounded.png" style="display: block; height: auto; border: 0; width: 600px; max-width: 100%;" width="600"/></div>
        </td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; background-color: #ffffff; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 10px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr><td class="pad" style="padding-bottom:5px;padding-left:20px;padding-right:20px;padding-top:5px;width:100%;">
        <div align="center" class="alignment" style="line-height:10px"><img class="big" src="https://i.ibb.co/C8S6tTm/adminision-bg.jpg" style="display: block; height: auto; border: 0; width: 560px; max-width: 100%;" width="560"/></div>
        </td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="heading_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:5px;padding-top:25px;text-align:center;width:100%;">
        <h1 style="margin: 0; color: #555555; direction: ltr; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; font-size: 36px; font-weight: 400; letter-spacing: normal; line-height: 120%; text-align: center; margin-top: 0; margin-bottom: 0;"><span class="tinyMce-placeholder">Request Confirmation</span></h1>
        </td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;">
        <div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 25.2px; color: #737487; line-height: 1.8; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
        <p style="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">Hi! '''
        
        body_1 = '''</span><br/><br/></p>
        <tp style="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">Good Day,</span></p>
        <p syle="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">This is to inform you that your Appointment has been Confirmed, Your Appointment is on (<b>'''
        
        suffix = '''</b>)</span></p>
        </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-11" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
            <tbody><tr><td>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
            <tbody><tr>
            <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
            <table border="0" cellpadding="0" cellspacing="0" class="text_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr>
            <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;"><div style="font-family: sans-serif">
            <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 25.2px; color: #737487; line-height: 1.8; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
            <p style="margin: 0; font-size: 18px; text-align: justify; mso-line-height-alt: 32.4px;"><span style="font-size:18px;color:#000000;">Please be reminded of the following:</span></p>
            </div></div></td></tr></table>
            <table border="0" cellpadding="10" cellspacing="0" class="list_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr><td class="pad">
            <ul style="margin: 0; padding: 0; margin-left: 20px; list-style-type: square; color: #000000; font-size: 16px; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; font-weight: 400; line-height: 150%; text-align: left; direction: ltr; letter-spacing: 0px;">
            <li style="list-style-type: none;"><ul style="margin: 0; padding: 0; list-style-type: square; margin-top: 4px;">
            <li style="margin-bottom: 4px; margin-left: 23px;">Please bring your own ballpen</li>
            <li style="margin-bottom: 4px; margin-left: 23px;">No facemask No face shield No Entry </li>
            <li style="margin-bottom: 4px; margin-left: 23px;">Bring 1 valid ID with present address and a photocopy</li>
            <li style="margin-bottom: 4px; margin-left: 23px;">Please wear proper dress (non-revealing dress) both for male & female examinees. Dress modestly so that you will be allowed to enter the premises of UDM</li>
            </ul></li></ul></td></tr></table>
            <table border="0" cellpadding="0" cellspacing="0" class="text_block block-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr><td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:30px;"><div style="font-family: sans-serif">
            <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 28px; color: #737487; line-height: 2; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
            <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;"><span style="color:#000000;">Failure to comply to this reminders will be ignored and will not be entertained.</span></span></p>
            <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;"><span style="color:#000000;">&nbsp;</span></span></p>
            <p style="margin: 0; font-size: 17px; text-align: center; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;color:#000000;"><b>This confirmation of appointment is for confirmation of your academic details in our university. In order to process your request in out university</b></span></p>
            <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;"><span style="color:#000000;">&nbsp;</span></span></p>
            <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;color:#000000;">See you at the campus,</span></p>
            <p style="margin: 0; font-size: 17px; text-align: justify; letter-spacing: normal;"><strong><span style="font-size:17px;color:#000000;">UDM Registrar Office</span></strong></p>
            </div></div></td></tr></table>
            <table border="0" cellpadding="0" cellspacing="0" class="text_block block-4" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr>
            <td class="pad" style="padding-left:15px;padding-right:15px;padding-top:20px;"><div style="font-family: sans-serif">
            <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 28px; color: #737487; line-height: 2; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
            <p style="margin: 0; font-size: 17px; text-align: center; letter-spacing: normal;"><span style="color:#000000;">This is a auto response, <b>DO NOT REPLY TO THIS EMAIL</b></span></p>
            </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-12" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
            <tbody><tr><td>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
            <tbody><tr>
            <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
            <table border="0" cellpadding="0" cellspacing="0" class="button_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
            <tr>
            <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;text-align:center;"><div align="center" class="alignment">
            <!--[if mso]><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://udm.edu.ph/udm2/" style="height:52px;width:236px;v-text-anchor:middle;" arcsize="8%" stroke="false" fillcolor="#4f9f21"><w:anchorlock/><v:textbox inset="0px,0px,0px,0px"><center style="color:#ffffff; font-family:Arial, sans-serif; font-size:16px"><![endif]--><a href="https://udm.edu.ph/udm2/" style="text-decoration:none;display:inline-block;color:#ffffff;background-color:#4f9f21;border-radius:4px;width:auto;border-top:1px solid #4f9f21;font-weight:400;border-right:1px solid #4f9f21;border-bottom:1px solid #4f9f21;border-left:1px solid #4f9f21;padding-top:10px;padding-bottom:10px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;text-align:center;mso-border-alt:none;word-break:keep-all;" target="_blank"><span style="padding-left:60px;padding-right:60px;font-size:16px;display:inline-block;letter-spacing:normal;"><span dir="ltr" style="word-break: break-word; line-height: 32px;"><em>Visit our website</em></span></span></a>
            <!--[if mso]></center></v:textbox></v:roundrect><![endif]--></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-13" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
            <tbody><tr><td>
            <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; width: 600px;" width="600">
            <tbody><tr>
            <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
            <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
            <tr><td class="pad" style="width:100%;padding-right:0px;padding-left:0px;">
            <div align="center" class="alignment" style="line-height:10px"><img class="fullMobileWidth" src="https://d1oco4z2z1fhwp.cloudfront.net/templates/default/4011/bottom-rounded.png" style="display: block; height: auto; border: 0; width: 600px; max-width: 100%;" width="600"/></div>
            </td></tr></table>
            <table border="0" cellpadding="0" cellspacing="0" class="text_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr>
            <td class="pad" style="padding-bottom:5px;padding-left:5px;padding-right:5px;padding-top:30px;"><div style="font-family: Georgia, 'Times New Roman', serif">
            <div class="txtTinyMce-wrapper" style="font-size: 12px; font-family: Georgia, Times, 'Times New Roman', serif; mso-line-height-alt: 14.399999999999999px; color: #262b30; line-height: 1.2;">
            <p style="margin: 0; font-size: 14px; text-align: center;"><span style="color:#ffffff;font-size:18px;"><em>Public service through quality education</em></span></p>
            </div></div></td></tr></table>
            <table border="0" cellpadding="0" cellspacing="0" class="text_block block-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
            <tr><td class="pad" style="padding-bottom:35px;padding-left:10px;padding-right:10px;padding-top:5px;"><div style="font-family: sans-serif">
            <div class="txtTinyMce-wrapper" style="font-size: 12px; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; mso-line-height-alt: 14.399999999999999px; color: #262b30; line-height: 1.2;">
            <p style="margin: 0; text-align: center; font-size: 13px;"><span style="font-size:13px;">Justice Cecilia Munoz-Palma corner Mayor Antonio J. Villegas Street Mehan Gardens, Ermita, Manila, Philippines 1000</span></p>
            <p style="margin: 0; text-align: center; font-size: 13px; mso-line-height-alt: 14.399999999999999px;"> </p>
            <p style="margin: 0; text-align: center; font-size: 13px;"><span style="font-size:13px;">Tel. No.: (+632)5336-6582 / (+632)5336-8956 / (+632)5336-8966 | Fax. No. (+632)336-6554 | www.udm.edu.ph</span></p>
            </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table></body></html>'''
        
        msg.set_content("%s <b>%s</b>, %s %s %s" % (prefix, name, body_1, scheduled_datetime, suffix), 'html')
        with smtplib.SMTP_SSL(gmail_host_smtplib, 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)

    def notify(self):
        name = self.entry_1.get()
        course = self.combobox_1.get()
        section = self.combobox_2.get()
        request_type = self.combobox_4.get()
        email = self.entry_2.get()
        student_no = self.entry_3.get()
        return_date = self.cal.get_date().strftime("%B %d, %Y") + " " + self.combobox_5.get()
        regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        if course == "Section Select" or section == "Section Select" or request_type == "Request Type Select" or name == "" or student_no == "":
            ctypes.windll.user32.MessageBoxW(0, "Blank box detected, Please fill all boxes", "Attention", MB_OK | ICON_EXLAIM)
            return
        if (re.fullmatch(regex, email)) == None:
            ctypes.windll.user32.MessageBoxW(0, "This is not an email, please check the spelling", "Attention", MB_OK | ICON_EXLAIM)
            return
        if self.combobox_5.get() == "Time Select":
            ctypes.windll.user32.MessageBoxW(0, "Please select desired time first", "Attention", MB_OK | ICON_EXLAIM)
            return
        msg = EmailMessage()
        msg['Subject'] = 'Universidad De Manila - Update for your request'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = email
        prefix = '''<!DOCTYPE html>
        <html lang="en" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:v="urn:schemas-microsoft-com:vml">
        <head>
        <title></title>
        <meta content="text/html; charset=utf-8" http-equiv="Content-Type"/>
        <meta content="width=device-width, initial-scale=1.0" name="viewport"/>
        <!--[if mso]><xml><o:OfficeDocumentSettings><o:PixelsPerInch>96</o:PixelsPerInch><o:AllowPNG/></o:OfficeDocumentSettings></xml><![endif]-->
        <style> 
                * {
                    box-sizing: border-box;
                }
                body {
                    margin: 0;
                    padding: 0;
                }
                a[x-apple-data-detectors] {
                    color: inherit !important;
                    text-decoration: inherit !important;
                }
                #MessageViewBody a {
                    color: inherit;
                    text-decoration: none;
                }
                p {
                    line-height: inherit
                }
                .desktop_hide,
                .desktop_hide table {
                    mso-hide: all;
                    display: none;
                    max-height: 0px;
                    overflow: hidden;
                }
                @media (max-width:620px) {
                    .desktop_hide table.icons-inner {
                        display: inline-block !important;
                    }
                    .icons-inner {
                        text-align: center;
                    }
                    .icons-inner td {
                        margin: 0 auto;
                    }
                    .fullMobileWidth,
                    .image_block img.big,
                    .row-content {
                        width: 100% !important;
                    }
                    .mobile_hide {
                        display: none;
                    }
                    .stack .column {
                        width: 100%;
                        display: block;
                    }
                    .mobile_hide {
                        min-height: 0;
                        max-height: 0;
                        max-width: 0;
                        overflow: hidden;
                        font-size: 0px;
                    }
                    .desktop_hide,
                    .desktop_hide table {
                        display: table !important;
                        max-height: none !important;
                    }
                }
            </style>
        </head>
        <body style="background-color: #FFFFFF; margin: 0; padding: 0; -webkit-text-size-adjust: none; text-size-adjust: none;">
        <table border="0" cellpadding="0" cellspacing="0" class="nl-container" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #FFFFFF;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; background-image: url('https://i.ibb.co/LptDwSH/blue-glow-3-2.png'); background-repeat: no-repeat; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:35px;padding-left:30px;padding-right:30px;padding-top:35px;width:100%;">
        <div align="center" class="alignment" style="line-height:10px"><img src="https://i.ibb.co/JtmhRLG/icon.png" style="display: block; height: auto; border: 0; width: 150px; max-width: 100%;" width="150"/></div>
        </td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="width:100%;padding-right:0px;padding-left:0px;">
        <div align="center" class="alignment" style="line-height:10px"><img class="fullMobileWidth" src="https://d1oco4z2z1fhwp.cloudfront.net/templates/default/4011/top-rounded.png" style="display: block; height: auto; border: 0; width: 600px; max-width: 100%;" width="600"/></div>
        </td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; background-color: #ffffff; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 10px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr><td class="pad" style="padding-bottom:5px;padding-left:20px;padding-right:20px;padding-top:5px;width:100%;">
        <div align="center" class="alignment" style="line-height:10px"><img class="big" src="https://i.ibb.co/C8S6tTm/adminision-bg.jpg" style="display: block; height: auto; border: 0; width: 560px; max-width: 100%;" width="560"/></div>
        </td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="heading_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:5px;padding-top:25px;text-align:center;width:100%;">
        <h1 style="margin: 0; color: #555555; direction: ltr; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; font-size: 36px; font-weight: 400; letter-spacing: normal; line-height: 120%; text-align: center; margin-top: 0; margin-bottom: 0;"><span class="tinyMce-placeholder">Request Confirmation</span></h1>
        </td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;">
        <div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 25.2px; color: #737487; line-height: 1.8; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
        <p style="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">Hi! '''
        
        body_1 = '''</span><br/><br/></p>
        <tp style="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">Good Day,</span></p>
        <p syle="margin: 0; font-size: 17px; mso-line-height-alt: 30.6px;"><span style="font-size:17px;color:#000000;">This is to inform you that your '''
        
        body_2 = ''' has been processed and is ready to be released. Please be minded of the following details containing the release date:</span></p>
        </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-4" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr><td class="pad" style="padding-bottom:15px;padding-left:10px;padding-right:10px;padding-top:15px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Name: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''
        
        body_3 = '''</p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-5" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:15px;padding-left:10px;padding-right:10px;padding-top:15px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Type of Request: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''
        
        body_4 = ''' </p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-7" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%"><tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Course: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr><td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''
            
        body_5 = '''</p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-8" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Section: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''

        body_6 = '''</p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-9" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Student Number: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''
            
        body_7 = '''</p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-10" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; border-radius: 0; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">Date and Time of Release: </p></div></td></tr></table></td>
        <td class="column column-2" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="50%">
        <table border="0" cellpadding="0" cellspacing="0" class="paragraph_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-top:15px;padding-right:10px;padding-bottom:15px;padding-left:10px;">
        <div style="color:#000000;font-size:16px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;font-weight:400;line-height:120%;text-align:center;direction:ltr;letter-spacing:0px;mso-line-height-alt:19.2px;">
        <p style="margin: 0;">'''
        
        suffix = '''</p></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-11" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;"><div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 25.2px; color: #737487; line-height: 1.8; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
        <p style="margin: 0; font-size: 18px; text-align: justify; mso-line-height-alt: 32.4px;"><span style="font-size:18px;color:#000000;">Please be reminded of the following:</span></p>
        </div></div></td></tr></table>
        <table border="0" cellpadding="10" cellspacing="0" class="list_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr><td class="pad">
        <ul style="margin: 0; padding: 0; margin-left: 20px; list-style-type: square; color: #000000; font-size: 16px; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; font-weight: 400; line-height: 150%; text-align: left; direction: ltr; letter-spacing: 0px;">
        <li style="list-style-type: none;"><ul style="margin: 0; padding: 0; list-style-type: square; margin-top: 4px;">
        <li style="margin-bottom: 4px; margin-left: 23px;">Attached attachment contain your claim stub which will be needed for release. Releasing without the claiming stub will no be entertained</li>
        <li style="margin-bottom: 4px; margin-left: 23px;">Only visit the campus pertaining for your release date not before or after the dictated release date. </li>
        <li style="margin-bottom: 4px; margin-left: 23px;">Bring consent from the student if relatives or guardian will be the one who will received the request</li>
        <li style="margin-bottom: 4px; margin-left: 23px;">Bring your face mask and alcohol</li>
        <li style="margin-bottom: 4px; margin-left: 23px;">Please wear proper dress (non-revealing dress) both for male & female examinees. Dress modestly so that you will be allowed to enter the premises of UDM</li>
        </ul></li></ul></td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr><td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:30px;"><div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 28px; color: #737487; line-height: 2; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
        <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;"><span style="color:#000000;">Failure to comply to this reminders will be ignored and will not be entertained</span></span></p>
        <p style="margin: 0; font-size: 17px; text-align: justify; mso-line-height-alt: 34px; letter-spacing: normal;"><span style="font-size:17px;color:#000000;">See you at the campus,</span></p>
        <p style="margin: 0; font-size: 17px; text-align: justify; letter-spacing: normal;"><strong><span style="font-size:17px;color:#000000;">UDM Registrar Office</span></strong></p>
        </div></div></td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-4" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-left:15px;padding-right:15px;padding-top:20px;"><div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 14px; mso-line-height-alt: 28px; color: #737487; line-height: 2; font-family: Arial, Helvetica Neue, Helvetica, sans-serif;">
        <p style="margin: 0; font-size: 17px; text-align: center; letter-spacing: normal;"><span style="color:#000000;">This is a auto response, <b>DO NOT REPLY TO THIS EMAIL</b></span></p>
        </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-12" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21; background-size: auto;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #ffffff; color: #000000; background-size: auto; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="button_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:20px;padding-left:15px;padding-right:15px;padding-top:20px;text-align:center;"><div align="center" class="alignment">
        <!--[if mso]><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://udm.edu.ph/udm2/" style="height:52px;width:236px;v-text-anchor:middle;" arcsize="8%" stroke="false" fillcolor="#4f9f21"><w:anchorlock/><v:textbox inset="0px,0px,0px,0px"><center style="color:#ffffff; font-family:Arial, sans-serif; font-size:16px"><![endif]--><a href="https://udm.edu.ph/udm2/" style="text-decoration:none;display:inline-block;color:#ffffff;background-color:#4f9f21;border-radius:4px;width:auto;border-top:1px solid #4f9f21;font-weight:400;border-right:1px solid #4f9f21;border-bottom:1px solid #4f9f21;border-left:1px solid #4f9f21;padding-top:10px;padding-bottom:10px;font-family:Arial, Helvetica Neue, Helvetica, sans-serif;text-align:center;mso-border-alt:none;word-break:keep-all;" target="_blank"><span style="padding-left:60px;padding-right:60px;font-size:16px;display:inline-block;letter-spacing:normal;"><span dir="ltr" style="word-break: break-word; line-height: 32px;"><em>Visit our website</em></span></span></a>
        <!--[if mso]></center></v:textbox></v:roundrect><![endif]--></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row row-13" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #4f9f21;" width="100%">
        <tbody><tr><td>
        <table align="center" border="0" cellpadding="0" cellspacing="0" class="row-content stack" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-position: center top; color: #000000; width: 600px;" width="600">
        <tbody><tr>
        <td class="column column-1" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; font-weight: 400; text-align: left; vertical-align: top; padding-top: 0px; padding-bottom: 0px; border-top: 0px; border-right: 0px; border-bottom: 0px; border-left: 0px;" width="100%">
        <table border="0" cellpadding="0" cellspacing="0" class="image_block block-1" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt;" width="100%">
        <tr><td class="pad" style="width:100%;padding-right:0px;padding-left:0px;">
        <div align="center" class="alignment" style="line-height:10px"><img class="fullMobileWidth" src="https://d1oco4z2z1fhwp.cloudfront.net/templates/default/4011/bottom-rounded.png" style="display: block; height: auto; border: 0; width: 600px; max-width: 100%;" width="600"/></div>
        </td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-2" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr>
        <td class="pad" style="padding-bottom:5px;padding-left:5px;padding-right:5px;padding-top:30px;"><div style="font-family: Georgia, 'Times New Roman', serif">
        <div class="txtTinyMce-wrapper" style="font-size: 12px; font-family: Georgia, Times, 'Times New Roman', serif; mso-line-height-alt: 14.399999999999999px; color: #262b30; line-height: 1.2;">
        <p style="margin: 0; font-size: 14px; text-align: center;"><span style="color:#ffffff;font-size:18px;"><em>Public service through quality education</em></span></p>
        </div></div></td></tr></table>
        <table border="0" cellpadding="0" cellspacing="0" class="text_block block-3" role="presentation" style="mso-table-lspace: 0pt; mso-table-rspace: 0pt; word-break: break-word;" width="100%">
        <tr><td class="pad" style="padding-bottom:35px;padding-left:10px;padding-right:10px;padding-top:5px;"><div style="font-family: sans-serif">
        <div class="txtTinyMce-wrapper" style="font-size: 12px; font-family: Arial, Helvetica Neue, Helvetica, sans-serif; mso-line-height-alt: 14.399999999999999px; color: #262b30; line-height: 1.2;">
        <p style="margin: 0; text-align: center; font-size: 13px;"><span style="font-size:13px;">Justice Cecilia Munoz-Palma corner Mayor Antonio J. Villegas Street Mehan Gardens, Ermita, Manila, Philippines 1000</span></p>
        <p style="margin: 0; text-align: center; font-size: 13px; mso-line-height-alt: 14.399999999999999px;"> </p>
        <p style="margin: 0; text-align: center; font-size: 13px;"><span style="font-size:13px;">Tel. No.: (+632)5336-6582 / (+632)5336-8956 / (+632)5336-8966 | Fax. No. (+632)336-6554 | www.udm.edu.ph</span></p>
        </div></div></td></tr></table></td></tr></tbody></table></td></tr></tbody></table></body></html>'''
        
        msg.set_content("%s <b>%s</b>, %s %s%s%s%s%s%s%s%s%s%s%s%s%s%s" % (prefix, name, body_1, request_type,
                                                                       body_2, name, body_3,  request_type, 
                                                                       body_4, course, body_5, section, 
                                                                       body_6, student_no, body_7, return_date,
                                                                       suffix), 'html')
        with smtplib.SMTP_SSL(gmail_host_smtplib, 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)

if __name__ == "__main__":
    MB_OK = 0x0
    #MB_OKCXL = 0x01
    #MB_YESNOCXL = 0x03
    #MB_YESNO = 0x04
    #MB_HELP = 0x4000
    ICON_EXLAIM=0x30
    #ICON_INFO = 0x40
    ICON_STOP = 0x10
    app = App()
    app.start()

